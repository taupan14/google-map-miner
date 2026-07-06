import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any
import io


def generate_xlsx(job: Dict[str, Any], places: List[Dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # ── Color palette ─────────────────────────────────────────
    COLOR_HEADER_BG = "0F172A"
    COLOR_HEADER_FG = "F8FAFC"
    COLOR_ACCENT = "6366F1"
    COLOR_ROW_ALT = "F1F5F9"
    COLOR_BORDER = "CBD5E1"

    # ── Job Info Header ───────────────────────────────────────
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"Map Miner Export — {job['keyword']}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=COLOR_HEADER_FG)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # Meta row
    ws.merge_cells("A2:K2")
    location = " | ".join(filter(None, [
        job.get("country"), job.get("province"),
        job.get("city"), job.get("district")
    ]))
    meta_parts = [
        f"Location: {location}",
        f"Total: {len(places)} places",
        f"Status: {job.get('status', '').upper()}",
        f"Exported: {datetime.now().strftime('%d %b %Y %H:%M')}",
    ]
    meta_cell = ws["A2"]
    meta_cell.value = "   •   ".join(meta_parts)
    meta_cell.font = Font(name="Calibri", size=10, color="64748B")
    meta_cell.fill = PatternFill("solid", fgColor="1E293B")
    meta_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    # Spacer
    ws.row_dimensions[3].height = 6

    # ── Column Headers ────────────────────────────────────────
    headers = [
        ("No", 5),
        ("Place Name", 35),
        ("Category", 20),
        ("Rating", 8),
        ("Reviews", 10),
        ("Address", 40),
        ("Phone", 18),
        ("Website", 30),
        ("Latitude", 14),
        ("Longitude", 14),
        ("Google Maps URL", 50),
    ]

    header_row = 4
    header_fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    header_font = Font(name="Calibri", bold=True, size=10, color=COLOR_HEADER_FG)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 22

    # ── Data Rows ─────────────────────────────────────────────
    thin_border = Border(
        left=Side(style="thin", color=COLOR_BORDER),
        right=Side(style="thin", color=COLOR_BORDER),
        top=Side(style="thin", color=COLOR_BORDER),
        bottom=Side(style="thin", color=COLOR_BORDER),
    )

    data_font = Font(name="Calibri", size=10)
    alt_fill = PatternFill("solid", fgColor=COLOR_ROW_ALT)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for row_idx, place in enumerate(places, start=1):
        row_num = header_row + row_idx
        is_alt = row_idx % 2 == 0
        row_fill = alt_fill if is_alt else None

        row_data = [
            row_idx,
            place.get("place_name", ""),
            place.get("category", ""),
            place.get("rating", ""),
            place.get("reviews", ""),
            place.get("address", ""),
            place.get("phone", ""),
            place.get("website", ""),
            place.get("latitude", ""),
            place.get("longitude", ""),
            place.get("maps_url", ""),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

            # Alignment per column type
            if col_idx in (1, 4, 5, 9, 10):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

            # Hyperlink for maps_url
            if col_idx == 11 and value and str(value).startswith("http"):
                cell.hyperlink = str(value)
                cell.font = Font(name="Calibri", size=10, color="6366F1", underline="single")

        ws.row_dimensions[row_num].height = 18

    # ── Freeze panes ──────────────────────────────────────────
    ws.freeze_panes = f"A{header_row + 1}"

    # ── Auto filter ───────────────────────────────────────────
    ws.auto_filter.ref = f"A{header_row}:K{header_row + len(places)}"

    # ── Summary Sheet ─────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 30

    summary_title = ws_summary["A1"]
    ws_summary.merge_cells("A1:B1")
    summary_title.value = "Job Summary"
    summary_title.font = Font(name="Calibri", bold=True, size=13, color=COLOR_HEADER_FG)
    summary_title.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    summary_title.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 28

    summary_data = [
        ("Keyword", job.get("keyword", "")),
        ("Country", job.get("country", "")),
        ("Province", job.get("province", "") or "-"),
        ("City", job.get("city", "") or "-"),
        ("District", job.get("district", "") or "-"),
        ("Status", job.get("status", "").upper()),
        ("Total Found", job.get("total_found", 0)),
        ("Total Scraped", job.get("total_scraped", 0)),
        ("Places in Export", len(places)),
        ("Started At", job.get("started_at", "-")),
        ("Finished At", job.get("finished_at", "-")),
        ("Exported At", datetime.now().strftime("%d %b %Y %H:%M:%S")),
    ]

    key_font = Font(name="Calibri", bold=True, size=10, color="334155")
    val_font = Font(name="Calibri", size=10, color="1E293B")

    for i, (key, val) in enumerate(summary_data, start=2):
        cell_key = ws_summary.cell(row=i, column=1, value=key)
        cell_val = ws_summary.cell(row=i, column=2, value=str(val) if val else "-")
        cell_key.font = key_font
        cell_val.font = val_font
        if i % 2 == 0:
            cell_key.fill = PatternFill("solid", fgColor="F1F5F9")
            cell_val.fill = PatternFill("solid", fgColor="F1F5F9")
        ws_summary.row_dimensions[i].height = 18

    # ── Save to buffer ────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
